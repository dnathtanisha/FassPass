import cv2
import os
import time
from openpyxl import Workbook, load_workbook

DATASET_DIR = "datasets"
EXCEL_FILE = "user_details.xlsx"
POSES = ["center", "left", "right", "up", "down"]

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

def register_user():
    if not os.path.exists(DATASET_DIR):
        os.makedirs(DATASET_DIR)

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        print("Error: Could not open webcam.")
        return

    name = input("Enter your name: ").strip()
    college = input("Enter your College/University: ").strip()

    for pose in POSES:
        print(f"Turn your face to the {pose}. Capturing in 5 seconds...")
        start_time = time.time()
        captured = False

        while not captured and (time.time() - start_time < 7):
            ret, frame = cap.read()
            if not ret:
                break
            frame = cv2.flip(frame, 1)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)

            cv2.putText(frame, f"Pose: {pose}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.imshow("Registration", frame)

            if time.time() - start_time >= 5 and len(faces) > 0:
                (x, y, w, h) = max(faces, key=lambda rect: rect[2] * rect[3])
                face_img = frame[y:y + h, x:x + w]
                filename = f"{name.replace(' ', '')}{college.replace(' ', '')}{pose}_{int(time.time())}.jpg"
                cv2.imwrite(os.path.join(DATASET_DIR, filename), face_img)
                print(f"Saved: {filename}")
                captured = True

            if cv2.waitKey(1) & 0xFF == ord('q'):
                print("Registration cancelled.")
                cap.release()
                cv2.destroyAllWindows()
                return

    # Save to Excel
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Name", "College/University"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active
    sheet.append([name, college])
    wb.save(EXCEL_FILE)

    cap.release()
    cv2.destroyAllWindows()
    print(f"Registration complete for {name}")

if __name__ == "__main__":
    register_user()
