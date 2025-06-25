import streamlit as st
import cv2
import numpy as np
import os
import re
import time
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk # PIL is still used for image manipulation/conversion if needed
import threading
import queue # For thread-safe communication between VideoProcessor and main Streamlit thread
from streamlit_webrtc import webrtc_streamer, VideoProcessorBase # Import here
import av # Needed by streamlit-webrtc for video frames

# --- Backend Helper Functions ---

def _setup_excel_file(excel_file):
    """Setup Excel file with headers if it doesn't exist."""
    try:
        if not os.path.exists(excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "College/University"])
            workbook.save(excel_file)
        return True
    except Exception as e:
        print(f"Error setting up Excel file: {str(e)}")
        return False

def _save_user_details(excel_file, name, college):
    """Save user details to Excel file."""
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        # Check if user already exists
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == college:
                print(f"User {name} from {college} already exists in Excel.")
                return True # Consider it successful if already exists
        sheet.append([name, college])
        workbook.save(excel_file)
        return True
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        return False

# --- Face Recognition Backend Logic (Adapted for Streamlit) ---

class FaceRecognizer:
    """An improved face recognition system with multiple matching algorithms."""
    
    def __init__(self, dataset_dir="datasets", recognition_threshold=0.6, min_votes=1):
        """
        Initialize the Face Recognizer with improved algorithms.
        
        Args:
            dataset_dir (str): Directory containing face images
            recognition_threshold (float): Threshold for face recognition (0.0-1.0)
            min_votes (int): Minimum votes needed for recognition
        """
        self.dataset_dir = dataset_dir
        self.recognition_threshold = recognition_threshold
        self.min_votes = min_votes
        # Ensure 'haarcascade_frontalface_default.xml' is in the same directory or accessible
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        if self.face_cascade.empty():
            st.error("ERROR: Face cascade not loaded. Check 'haarcascade_frontalface_default.xml' path in your project directory.")
        
        self.known_faces = defaultdict(list)  # {person_id: [face_images]} 
        self.person_info = {}  # {person_id: (name, college)}
        
        # Initialize face recognizers
        self.lbph_recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.eigen_recognizer = cv2.face.EigenFaceRecognizer_create()
        self.fisher_recognizer = cv2.face.FisherFaceRecognizer_create()
        
        self.recognizers_trained = False
        self.label_to_person = {} # Map integer labels from recognizers back to person_id

    def preprocess_face(self, face_img, target_size=(100, 100)):
        """
        Preprocess face image for better recognition.
        
        Args:
            face_img (numpy.ndarray): Input face image (grayscale)
            target_size (tuple): Target size for resizing
            
        Returns:
            numpy.ndarray: Preprocessed face image
        """
        if face_img.ndim == 3: # Convert to grayscale if it's not already
            face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2GRAY)

        # Resize to standard size
        face_img = cv2.resize(face_img, target_size, interpolation=cv2.INTER_AREA)
        
        # Histogram equalization for better contrast
        face_img = cv2.equalizeHist(face_img)
        
        # Gaussian blur to reduce noise
        face_img = cv2.GaussianBlur(face_img, (3, 3), 0)
        
        # Normalize pixel values
        face_img = cv2.normalize(face_img, None, 0, 255, cv2.NORM_MINMAX)
        
        return face_img
    
    def calculate_similarity_metrics(self, face1, face2):
        """
        Calculate multiple similarity metrics between two preprocessed faces.
        
        Args:
            face1, face2 (numpy.ndarray): Preprocessed face images to compare
            
        Returns:
            dict: Dictionary of similarity scores
        """
        # Ensure images are float32 for comparison where necessary
        face1_f = face1.astype(np.float32)
        face2_f = face2.astype(np.float32)

        # Template matching (normalized correlation)
        result = cv2.matchTemplate(face1, face2, cv2.TM_CCOEFF_NORMED)
        template_score = np.max(result)
        
        # Histogram comparison (Correlation)
        hist1 = cv2.calcHist([face1], [0], None, [256], [0, 256])
        hist2 = cv2.calcHist([face2], [0], None, [256], [0, 256])
        hist_score = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
        
        # Structural similarity (simplified PSNR for image quality difference)
        mse = np.mean((face1_f - face2_f) ** 2)
        if mse == 0: # Images are identical
            psnr_score = 1.0
        else:
            max_pixel_value = 255.0
            psnr = 20 * np.log10(max_pixel_value / np.sqrt(mse))
            psnr_score = max(0.0, min(psnr / 60, 1.0)) # Normalize PSNR to a 0-1 scale (approx 60dB is max for 8-bit)
        
        return {
            'template': template_score,
            'histogram': hist_score,
            'psnr': psnr_score
        }
    
    def load_known_faces(self):
        """
        Load all known faces from the dataset directory with improved organization.
        
        Returns:
            bool: True if faces loaded successfully, False otherwise
        """
        if not os.path.exists(self.dataset_dir):
            print(f"Dataset directory '{self.dataset_dir}' does not exist. No faces to load.")
            return False
        
        self.known_faces.clear()
        self.person_info.clear()
        
        face_count = 0
        for filename in os.listdir(self.dataset_dir):
            if filename.endswith(".jpg"):
                # Extract name and college from filename (e.g., "JohnDoeSomeCollege_center_12345.jpg")
                match = re.match(r"([A-Za-z0-9]+)([A-Za-z0-9]+)_[a-z]+_\d+\.jpg", filename)
                if match:
                    name_part = match.group(1)
                    college_part = match.group(2)
                    person_id = f"{name_part}_{college_part}" # Unique ID for each person based on their name and college
                    
                    img_path = os.path.join(self.dataset_dir, filename)
                    
                    try:
                        img = cv2.imread(img_path)
                        if img is None:
                            print(f"Could not load image: {filename}")
                            continue
                            
                        # Convert to grayscale first for face detection
                        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                        faces = self.face_cascade.detectMultiScale(gray, 1.3, 5, minSize=(100, 100)) # Use a minSize for better detection
                        
                        if len(faces) > 0:
                            # Take the largest face in the image for registration
                            (x, y, w, h) = max(faces, key=lambda rect: rect[2] * rect[3])
                            face_img = gray[y:y + h, x:x + w] # Crop face directly from grayscale
                            face_img = self.preprocess_face(face_img) # Preprocess the cropped grayscale face
                            
                            self.known_faces[person_id].append(face_img)
                            self.person_info[person_id] = (name_part, college_part)
                            face_count += 1
                        else:
                            print(f"No dominant face detected in {filename} for training.")
                    except Exception as e:
                        print(f"Error processing {filename}: {str(e)}")
        
        print(f"Loaded {face_count} face images for {len(self.known_faces)} people.")
        
        # Train recognizers if we have enough data
        if len(self.known_faces) > 0:
            self.train_recognizers()
        
        return len(self.known_faces) > 0
    
    def train_recognizers(self):
        """Train the face recognizers with loaded data."""
        faces = []
        labels = []
        self.label_to_person.clear() # Clear previous mapping

        if len(self.known_faces) < 1: 
            self.recognizers_trained = False
            return
        
        current_label = 0
        for person_id, face_list in self.known_faces.items():
            if not face_list: 
                continue
            self.label_to_person[current_label] = person_id
            for face in face_list:
                faces.append(face)
                labels.append(current_label)
            current_label += 1
        
        if len(faces) < 1: 
            self.recognizers_trained = False
            return
        
        faces_np = np.array(faces)
        labels_np = np.array(labels)
        
        try:
            self.lbph_recognizer.train(faces_np, labels_np)
            print("LBPH recognizer trained.")
            
            if len(self.known_faces) >= 2 and all(len(f) > 0 for f in self.known_faces.values()):
                if len(faces_np) > 0 and all(f.shape == faces_np[0].shape for f in faces_np):
                    try:
                        self.eigen_recognizer.train(faces_np, labels_np)
                        print("Eigen recognizer trained.")
                    except cv2.error as e:
                        print(f"Warning: Could not train EigenFaceRecognizer. Error: {e}")
                    
                    try:
                        self.fisher_recognizer.train(faces_np, labels_np)
                        print("Fisher recognizer trained.")
                    except cv2.error as e:
                        print(f"Warning: Could not train FisherFaceRecognizer. Error: {e}")
                else:
                    print("Warning: Face image dimensions are not consistent for Eigen/Fisher training. Skipping.")
            else:
                print("Not enough unique people or samples per person for Eigen/Fisher training. Skipping.")
            
            self.recognizers_trained = True
            print("All available face recognizers training attempted.")
            
        except Exception as e:
            print(f"Error during recognizer training: {e}")
            self.recognizers_trained = False
    
    def recognize_face(self, face_img):
        """
        Recognize a face using multiple algorithms and voting.
        
        Args:
            face_img (numpy.ndarray): Face image to recognize (can be BGR or grayscale)
            
        Returns:
            tuple: (name, college, aggregated_confidence) or (None, None, 0.0) if not recognized
        """
        if len(self.known_faces) == 0 or not self.recognizers_trained:
            return None, None, 0.0
        
        processed_face_img = self.preprocess_face(face_img)
        votes = defaultdict(float) 
        
        try:
            label_lbph, confidence_lbph = self.lbph_recognizer.predict(processed_face_img)
            similarity_lbph = max(0.0, 1.0 - (confidence_lbph / 100.0)) 
            
            if similarity_lbph > self.recognition_threshold:
                person_id = self.label_to_person.get(label_lbph)
                if person_id:
                    votes[person_id] += similarity_lbph * 1.0 
        except Exception:
            pass 

        if self.eigen_recognizer and self.eigen_recognizer.getThreshold() >= 0: 
            try:
                label_eigen, confidence_eigen = self.eigen_recognizer.predict(processed_face_img)
                similarity_eigen = max(0.0, 1.0 - (confidence_eigen / 5000.0))
                if similarity_eigen > self.recognition_threshold:
                    person_id = self.label_to_person.get(label_eigen)
                    if person_id:
                        votes[person_id] += similarity_eigen * 0.5 
            except Exception:
                pass 
        
        if self.fisher_recognizer and self.fisher_recognizer.getThreshold() >= 0: 
            try:
                label_fisher, confidence_fisher = self.fisher_recognizer.predict(processed_face_img)
                similarity_fisher = max(0.0, 1.0 - (confidence_fisher / 3000.0))
                if similarity_fisher > self.recognition_threshold:
                    person_id = self.label_to_person.get(label_fisher)
                    if person_id:
                        votes[person_id] += similarity_fisher * 0.5 
            except Exception:
                pass

        for person_id_known, face_list_known in self.known_faces.items():
            best_combined_metric_score = 0.0
            
            for known_face in face_list_known:
                try:
                    metrics = self.calculate_similarity_metrics(processed_face_img, known_face)
                    combined_score = (
                        metrics['template'] * 0.4 +
                        metrics['histogram'] * 0.3 +
                        metrics['psnr'] * 0.3
                    )
                    best_combined_metric_score = max(best_combined_metric_score, combined_score)
                except Exception:
                    pass

            if best_combined_metric_score > self.recognition_threshold:
                votes[person_id_known] += best_combined_metric_score * 0.7 

        recognized_person_id = None
        highest_aggregated_confidence = 0.0

        for person_id, total_score in votes.items():
            if total_score > highest_aggregated_confidence:
                highest_aggregated_confidence = total_score
                recognized_person_id = person_id
        
        if recognized_person_id and highest_aggregated_confidence >= self.min_votes * self.recognition_threshold:
            name, college = self.person_info[recognized_person_id]
            return name, college, highest_aggregated_confidence
        
        return None, None, 0.0

# --- Streamlit Video Processors (for streamlit-webrtc) ---

class LiveRecognitionProcessor(VideoProcessorBase):
    def __init__(self, status_q):
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.recognizer = get_face_recognizer_cached() # Get the cached instance here
        self.status_q = status_q
        self.last_status_update = time.time()
        self.status_update_interval = 0.5 # seconds
        self.recognized_already = False # Flag to stop further processing after first recognition

        if self.face_cascade.empty():
            self.status_q.put("ERROR: Face cascade not loaded.")

    def recv_video_frame(self, frame):
        img = frame.to_ndarray(format="bgr24") # Convert WebRTC frame to OpenCV BGR
        
        # If already recognized, just return the last frame without further processing
        if self.recognized_already:
            flipped_img = cv2.flip(img, 1)
            return av.VideoFrame.from_ndarray(flipped_img, format="bgr24")

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(
            gray, 
            scaleFactor=1.1, 
            minNeighbors=5, 
            minSize=(50, 50),
            flags=cv2.CASCADE_SCALE_IMAGE
        )
        
        recognized_info_for_ui = "Unknown Face" # Default status if no face or not recognized
        text_color = (0, 0, 0) # Black text for drawing on image

        for (x, y, w, h) in faces:
            padding = 10
            x_pad = max(0, x - padding)
            y_pad = max(0, y - padding)
            w_pad = min(img.shape[1] - x_pad, w + 2 * padding)
            h_pad = min(img.shape[0] - y_pad, h + 2 * padding)
            
            face_img = gray[y_pad:y_pad + h_pad, x_pad:x_pad + w_pad]
            
            name, college, confidence = self.recognizer.recognize_face(face_img)
            
            if name and college:
                label = f"{name} ({college})"
                label += f" [{confidence:.2f}]"
                recognition_color = (0, 255, 0) # Green for recognized
                recognized_info_for_ui = f"Recognized: {name} ({college})"
                
                # Signal to stop recognition in main thread
                if not self.status_q.full():
                    self.status_q.put(("RECOGNIZED_AND_STOP", name, college))
                self.recognized_already = True # Set flag to stop further processing in this worker
                
            else:
                label = "Unknown"
                if confidence > 0: 
                    label += f" [{confidence:.2f}]"
                recognition_color = (0, 0, 255) # Red for unknown
            
            cv2.rectangle(img, (x, y), (x + w, y + h), recognition_color, 2)
            
            label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)[0]
            # Draw rectangle for text background (white box)
            cv2.rectangle(img, (x, y - 30), (x + label_size[0], y), recognition_color, -1)
            # Draw text (black color)
            cv2.putText(img, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, text_color, 2)
        
        # Update status queue only periodically to avoid flooding, unless it's a stop signal
        if not self.recognized_already and time.time() - self.last_status_update > self.status_update_interval:
            if not self.status_q.full():
                self.status_q.put(recognized_info_for_ui)
            self.last_status_update = time.time()

        # Flip the image horizontally as requested.
        flipped_img = cv2.flip(img, 1)

        return av.VideoFrame.from_ndarray(flipped_img, format="bgr24")

class RegistrationProcessor(VideoProcessorBase):
    def __init__(self, name, college, initial_pose_idx, status_q, countdown_q, progress_q):
        self.name = name
        self.college = college
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        self.poses = ["center", "left", "right", "up", "down"]
        self.timeout_per_pose = 5 # seconds to hold each pose
        self.images_per_pose = 15 # Number of images to capture for each pose (increase for robustness)

        self.current_pose_idx = initial_pose_idx
        self.current_pose_images = []
        self._capture_start_time = time.time() # Timer for current pose
        self.last_status_update_time = time.time()

        self.status_q = status_q
        self.countdown_q = countdown_q
        self.progress_q = progress_q
        self.recognizer = get_face_recognizer_cached() # Get the cached instance here
        
        if self.face_cascade.empty():
            self.status_q.put("ERROR: Face cascade not loaded.")

        if not os.path.exists("datasets"):
            os.makedirs("datasets")
        
        _setup_excel_file("user_details.xlsx")
        
        self.status_q.put(f"Starting registration. Please turn to {self.poses[self.current_pose_idx]}.")

    def recv_video_frame(self, frame):
        img = frame.to_ndarray(format="bgr24")
        flipped_img = cv2.flip(img, 1) # Flip for display
        display_img = flipped_img.copy()

        gray = cv2.cvtColor(flipped_img, cv2.COLOR_BGR2GRAY)
        faces = self.face_cascade.detectMultiScale(gray, 1.3, 5, minSize=(100, 100))
        
        for (x, y, w, h) in faces:
            cv2.rectangle(display_img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        
        # Update status and countdown messages periodically
        time_elapsed_pose = time.time() - self._capture_start_time
        remaining_time = max(0, int(self.timeout_per_pose - time_elapsed_pose))

        if time.time() - self.last_status_update_time > 0.5: # Update every 0.5 seconds
            if not self.countdown_q.full():
                self.countdown_q.put(remaining_time)
            if not self.status_q.full():
                current_pose_name = self.poses[self.current_pose_idx] if self.current_pose_idx < len(self.poses) else "Finished"
                self.status_q.put(f"Pose: {current_pose_name} - Hold for {remaining_time}s")
            self.last_status_update_time = time.time()

        # Capture logic: if a face is detected and timer for current pose is active
        if len(faces) > 0 and remaining_time < self.timeout_per_pose and self.current_pose_idx < len(self.poses):
            (x, y, w, h) = max(faces, key=lambda rect: rect[2] * rect[3]) # Take largest face
            face_image = gray[y:y + h, x:x + w] # Capture grayscale
            self.current_pose_images.append(face_image)
            
            # Draw number of captured images for current pose
            cv2.putText(display_img, f"Captured: {len(self.current_pose_images)}/{self.images_per_pose}", 
                        (img.shape[1] - 250, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
            
            # If enough images for current pose are captured OR timeout reached
            if len(self.current_pose_images) >= self.images_per_pose or remaining_time == 0:
                # Save images for this pose
                clean_name = re.sub(r'[^A-Za-z0-9]', '', self.name)
                clean_college = re.sub(r'[^A-Za-z0-9]', '', self.college)

                for idx, face_img_to_save in enumerate(self.current_pose_images):
                    filename = f"{clean_name}{clean_college}_{self.poses[self.current_pose_idx]}_{int(time.time())}_{idx}.jpg"
                    filepath = os.path.join("datasets", filename)
                    cv2.imwrite(filepath, face_img_to_save)
                
                self.current_pose_images = [] # Reset for next pose
                self.current_pose_idx += 1
                self._capture_start_time = time.time() # Reset timer for next pose

                total_progress_val = (self.current_pose_idx / len(self.poses)) * 100
                if not self.progress_q.full():
                    self.progress_q.put(total_progress_val)
                
                if self.current_pose_idx >= len(self.poses):
                    # All poses captured
                    _save_user_details("user_details.xlsx", self.name, self.college)
                    if not self.status_q.full():
                        self.status_q.put("Registration complete! Database updated. Click 'Stop'.")
                    if not self.countdown_q.full():
                        self.countdown_q.put(-1) # Signal completion
                    
                    # Call load_known_faces on the cached recognizer instance after new data is saved
                    # This ensures the recognition model is updated for new users
                    self.recognizer.load_known_faces()
                else:
                    if not self.status_q.full():
                        self.status_q.put(f"Pose captured. Next: {self.poses[self.current_pose_idx]}.")

        return av.VideoFrame.from_ndarray(display_img, format="bgr24")

# --- Streamlit UI Layout ---

st.set_page_config(page_title="AI Face System", layout="centered", initial_sidebar_state="auto")

# Function to get/create FaceRecognizer instance, managed by st.cache_resource
@st.cache_resource
def get_face_recognizer_cached():
    return FaceRecognizer()

# Initialize session state variables for persistence across reruns
if "recognition_status" not in st.session_state:
    st.session_state.recognition_status = "Please load database or start recognition."
if "reg_active" not in st.session_state:
    st.session_state.reg_active = False
if "reg_current_pose_idx" not in st.session_state:
    st.session_state.reg_current_pose_idx = 0
if "reg_name" not in st.session_state:
    st.session_state.reg_name = ""
if "reg_college" not in st.session_state:
    st.session_state.reg_college = ""
if "reg_status" not in st.session_state:
    st.session_state.reg_status = "Enter user details and click 'Start Registration'."
if "reg_countdown" not in st.session_state:
    st.session_state.reg_countdown = 5 # Initial countdown value
if "reg_progress_val" not in st.session_state:
    st.session_state.reg_progress_val = 0
if "rec_active" not in st.session_state:
    st.session_state.rec_active = False

# --- Queues for communication (global within the script context for simplicity) ---
# These queues will be passed to the VideoProcessors
reg_status_queue = queue.Queue(maxsize=1)
reg_countdown_queue = queue.Queue(maxsize=1)
reg_progress_queue = queue.Queue(maxsize=1)
rec_status_queue = queue.Queue(maxsize=1) # Max size 1 to ensure only latest status is kept
message_box_queue = queue.Queue(maxsize=1) # For general messages

st.title("AI Face Registration & Recognition System")

tab1, tab2 = st.tabs(["👤 Face Registration", "👀 Face Recognition"])

with tab1:
    st.header("Register New User")

    with st.form("registration_form"):
        reg_name_input = st.text_input("Name:", value=st.session_state.reg_name, key="reg_name_st")
        reg_college_input = st.text_input("College/University:", value=st.session_state.reg_college, key="reg_college_st")

        st.session_state.reg_name = reg_name_input
        st.session_state.reg_college = reg_college_input

        col_buttons = st.columns(2)
        with col_buttons[0]:
            submitted = st.form_submit_button("Start Registration", disabled=st.session_state.reg_active)
        with col_buttons[1]:
            stopped = st.form_submit_button("Stop Registration", disabled=not st.session_state.reg_active)
        
        if submitted and not st.session_state.reg_active:
            if not reg_name_input or not reg_college_input:
                message_box_queue.put(("Input Required", "Please enter both Name and College/University.", "warning"))
            else:
                st.session_state.reg_active = True
                st.session_state.reg_current_pose_idx = 0
                st.session_state.reg_progress_val = 0
                st.session_state.reg_status = "Registration started. Please follow instructions on camera feed."
                st.rerun() # Rerun to update UI and start camera
        elif stopped and st.session_state.reg_active:
            st.session_state.reg_active = False
            st.session_state.reg_current_pose_idx = 0 # Reset for next time
            st.session_state.reg_progress_val = 0
            st.session_state.reg_status = "Registration stopped by user."
            message_box_queue.put(("Stopped", "Face registration process has been stopped.", "info"))
            st.rerun()

    reg_status_area = st.empty()
    reg_countdown_area = st.empty()
    reg_progress_bar_area = st.empty()

    if st.session_state.reg_active:
        reg_status_area.info(st.session_state.reg_status)
        reg_countdown_area.markdown(f"<h3 style='text-align: center; color: #FF4500;'>Hold for {st.session_state.reg_countdown}s</h3>", unsafe_allow_html=True)
        reg_progress_bar_area.progress(st.session_state.reg_progress_val)
        
        ctx_reg = webrtc_streamer(
            key="registration_camera",
            video_processor_factory=lambda: RegistrationProcessor(
                name=st.session_state.reg_name,
                college=st.session_state.reg_college,
                initial_pose_idx=st.session_state.reg_current_pose_idx, # Pass current pose
                status_q=reg_status_queue,
                countdown_q=reg_countdown_queue,
                progress_q=reg_progress_queue,
            ),
            async_processing=True,
            media_stream_constraints={"video": True, "audio": False},
        )
        
        if ctx_reg.video_processor:
            pass 

with tab2:
    st.header("Recognize Faces")

    col_rec_buttons = st.columns(2)
    with col_rec_buttons[0]:
        if st.button("Load Database", key="load_db_button"):
            st.session_state.recognition_status = "Loading and training recognizer database..."
            if get_face_recognizer_cached().load_known_faces(): # Directly call load on cached instance
                st.session_state.recognition_status = f"Loaded {len(get_face_recognizer_cached().known_faces)} people. Recognizer trained."
                message_box_queue.put(("Database Loaded", "Face recognition database loaded and trained successfully.", "info"))
            else:
                st.session_state.recognition_status = "No known faces found. Please register faces first."
                message_box_queue.put(("No Data", "No known faces found in 'datasets' directory. Please register faces first.", "warning"))
            st.rerun() 
    with col_rec_buttons[1]:
        if st.button("Start Recognition", key="start_rec_button", disabled=st.session_state.reg_active): 
            st.session_state.recognition_status = "Starting recognition..."
            st.session_state.rec_active = True
            st.rerun()
        if st.button("Stop Recognition", key="stop_rec_button", disabled=not st.session_state.rec_active):
            st.session_state.rec_active = False
            st.session_state.recognition_status = "Recognition stopped."
            st.rerun()

    rec_status_area = st.empty()
    rec_status_area.info(st.session_state.recognition_status)

    if st.session_state.rec_active:
        ctx_rec = webrtc_streamer(
            key="recognition_camera",
            video_processor_factory=lambda: LiveRecognitionProcessor(
                status_q=rec_status_queue
            ),
            async_processing=True,
            media_stream_constraints={"video": True, "audio": False},
        )

# --- Periodic Queue Polling for UI Updates ---

def poll_queues_and_update_ui():
    try:
        reg_status_msg = reg_status_queue.get_nowait()
        st.session_state.reg_status = reg_status_msg
    except queue.Empty:
        pass

    try:
        reg_countdown_val = reg_countdown_queue.get_nowait()
        st.session_state.reg_countdown = reg_countdown_val
        if reg_countdown_val == -1: 
            st.session_state.reg_active = False
            st.session_state.reg_current_pose_idx = 0
            st.session_state.reg_progress_val = 100
            st.session_state.reg_status = "Registration process completed."
    except queue.Empty:
        pass
    
    try:
        reg_progress_val = reg_progress_queue.get_nowait()
        st.session_state.reg_progress_val = reg_progress_val
    except queue.Empty:
        pass

    try:
        rec_status_data = rec_status_queue.get_nowait()
        if isinstance(rec_status_data, tuple) and rec_status_data[0] == "RECOGNIZED_AND_STOP":
            _, name, college = rec_status_data
            st.session_state.recognition_status = f"Recognized: {name} ({college}). Recognition stopped automatically."
            st.session_state.rec_active = False # Stop the recognition camera
        else:
            st.session_state.recognition_status = rec_status_data
    except queue.Empty:
        pass

    try:
        title, message, icon_type = message_box_queue.get_nowait()
        if icon_type == "critical":
            st.error(f"Error: {message}") 
        elif icon_type == "warning":
            st.warning(f"Warning: {message}")
        elif icon_type == "info":
            st.info(f"Info: {message}")
    except queue.Empty:
        pass

poll_queues_and_update_ui()
 